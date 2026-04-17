"""
services/report_service.py
Generates Excel analytics report using pandas + openpyxl.
"""

import io
from datetime import datetime
from decimal import Decimal
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.models import Order
from config import get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL    = PatternFill("solid", fgColor="EBF0F8")


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_excel_report(orders: list[Order]) -> bytes:
    """
    Build a multi-sheet Excel report from a list of orders.
    Returns raw bytes to send as Telegram document.
    """
    if not orders:
        # Return empty workbook with note
        wb = _build_empty_wb()
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wb = _build_report_wb(orders)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_empty_wb():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws["A1"] = "Нет данных за выбранный период"
    return wb


def _build_report_wb(orders: list[Order]):
    from openpyxl import Workbook

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    _build_summary_sheet(ws_summary, orders)

    # ── Sheet 2: Orders list ──────────────────────────────────────────────
    ws_orders = wb.create_sheet("Заказы")
    _build_orders_sheet(ws_orders, orders)

    # ── Sheet 3: Top products ─────────────────────────────────────────────
    ws_products = wb.create_sheet("Топ продуктов")
    _build_products_sheet(ws_products, orders)

    return wb


def _write_header(ws, row: int, cols: list[str]) -> None:
    for col_idx, title in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _build_summary_sheet(ws, orders: list[Order]) -> None:
    total_revenue = sum(float(o.total_price) for o in orders)
    avg_check = total_revenue / len(orders) if orders else 0
    done_count = sum(1 for o in orders if o.status == "done")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    data = [
        ("📊 Отчёт сформирован", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("📋 Всего заказов", len(orders)),
        ("✅ Выполненных", done_count),
        ("💰 Выручка (тенге)", f"{total_revenue:,.0f}"),
        ("🧾 Средний чек (тенге)", f"{avg_check:,.0f}"),
    ]

    ws.cell(row=1, column=1, value="Ключевые показатели (последние 30 дней)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value).alignment = Alignment(horizontal="right")
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = ALT_FILL
            ws.cell(row=i, column=2).fill = ALT_FILL


def _build_orders_sheet(ws, orders: list[Order]) -> None:
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    headers = ["ID заказа", "Дата заказа", "Клиент", "Дата события", "Локация", "Статус", "Сумма (₸)"]
    _write_header(ws, 1, headers)
    ws.row_dimensions[1].height = 22

    for i, order in enumerate(orders, start=2):
        ws.cell(row=i, column=1, value=order.order_uid)
        ws.cell(row=i, column=2, value=order.created_at.strftime("%d.%m.%Y"))
        ws.cell(row=i, column=3, value=order.client_name)
        ws.cell(row=i, column=4, value=order.event_date.strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=i, column=5, value=order.location)
        ws.cell(row=i, column=6, value=order.status)
        ws.cell(row=i, column=7, value=float(order.total_price))
        ws.cell(row=i, column=7).number_format = '#,##0'

        if i % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = ALT_FILL

        for col in range(1, 8):
            ws.cell(row=i, column=col).border = _thin_border()


def _build_products_sheet(ws, orders: list[Order]) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18

    # Aggregate product data
    product_stats: dict[str, dict] = defaultdict(lambda: {"qty": 0, "revenue": 0.0})
    for order in orders:
        if order.items:
            for item in order.items:
                key = item.product_name
                product_stats[key]["qty"] += item.quantity
                product_stats[key]["revenue"] += item.quantity * float(item.unit_price)

    sorted_products = sorted(product_stats.items(), key=lambda x: x[1]["revenue"], reverse=True)

    headers = ["Продукт", "Кол-во продано", "Выручка (₸)"]
    _write_header(ws, 1, headers)
    ws.row_dimensions[1].height = 22

    for i, (name, stats) in enumerate(sorted_products, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=stats["qty"])
        ws.cell(row=i, column=3, value=stats["revenue"])
        ws.cell(row=i, column=3).number_format = '#,##0'

        if i % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=i, column=col).fill = ALT_FILL
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = _thin_border()
